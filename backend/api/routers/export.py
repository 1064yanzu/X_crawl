"""
数据导出路由
GET /api/v1/export/{task_id}?format=csv   - 导出为 CSV（UTF-8 BOM）
GET /api/v1/export/{task_id}?format=excel - 导出为 Excel（xlsx）
"""
import asyncio
import csv
import io
import json
import re
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Literal
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from api.services import task_db, task_manager

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/export", tags=["数据导出"])


# ─── 认证状态格式化辅助函数 ──────────────────────────────────────────────────

def _format_verified_status(author: dict, platform: str) -> str:
    """将认证字段格式化为可读的认证状态字符串。"""
    if platform == "youtube":
        return ""
    if platform == "weibo":
        verified = author.get("verified", False)
        v_type = author.get("verified_type", "")
        v_type_num = author.get("verified_type_num", -1)
        if not verified and v_type_num < 0:
            return ""
        if verified:
            if v_type in ("blue",) or v_type_num in (1, 2, 3):
                return "蓝V认证"
            if v_type in ("yellow",) or v_type_num == 0:
                return "黄V认证"
            return "已认证"
        return ""
    # X / Twitter
    blue = author.get("is_blue_verified", False)
    verified = author.get("verified", False)
    affiliate = author.get("affiliate_label", "")
    if affiliate:
        return f"关联标签: {affiliate}"
    if blue:
        return "蓝标认证"
    if verified:
        return "官方认证"
    return ""


def _format_verified_type(author: dict, platform: str) -> str:
    """将认证类型格式化为可读字符串。"""
    if platform == "youtube":
        return ""
    if platform == "weibo":
        v_type_num = author.get("verified_type_num", author.get("verified_type", -1))
        # 数字类型
        if isinstance(v_type_num, int):
            return {0: "个人", 1: "企业/机构", 2: "媒体", 3: "其他官方"}.get(v_type_num, "")
        # 文字类型（从 HTML 解析来的）
        if isinstance(v_type_num, str):
            return {"yellow": "个人", "blue": "企业/机构"}.get(v_type_num, v_type_num)
        return ""
    # X / Twitter
    parts = []
    v_type = author.get("verified_type", "")
    pro_type = author.get("professional_type", "")
    if v_type:
        parts.append(v_type)
    if pro_type and pro_type != v_type:
        parts.append(pro_type)
    return " / ".join(parts)

# 导出字段定义（顺序即为列顺序）
EXPORT_FIELDS = [
    # 基础
    ("platform", "平台"),
    ("id", "推文ID"),
    ("conversation_id", "对话ID"),
    ("created_at", "发布时间"),
    ("source", "发推客户端"),
    # 作者
    ("author_name", "作者昵称"),
    ("author_username", "作者账号"),
    ("author_id", "作者ID"),
    ("author_verified", "认证状态"),
    ("author_verified_type", "认证类型"),
    ("author_verified_reason", "认证说明"),
    ("author_professional", "专业账号"),
    ("author_followers", "作者粉丝数"),
    ("is_author", "是否楼主"),
    # 内容
    ("text", "推文内容"),
    ("lang", "语言"),
    # 互动指标
    ("like_count", "点赞数"),
    ("retweet_count", "转发数"),
    ("reply_count", "回复数"),
    ("quote_count", "引用数"),
    ("view_count", "浏览数"),
    ("bookmark_count", "收藏数"),
    # 链接
    ("url", "推文链接"),
    # 数据类型（区分原帖/回复）
    ("row_type", "数据类型"),
    ("parent_tweet_id", "所属推文ID"),
    # 回复关系
    ("reply_to_tweet_id", "回复目标推文ID"),
    ("reply_to_username", "回复目标用户"),
    # 类型标记
    ("is_retweet", "是否转推"),
    ("is_quote", "是否引用"),
    ("is_reply", "是否回复"),
    # 实体
    ("hashtags", "话题标签"),
    ("user_mentions_text", "提及用户"),
    # 媒体
    ("has_media", "含媒体"),
    ("media_types", "媒体类型"),
    ("media_urls", "媒体链接"),
]


def _flatten_tweet(tweet: dict) -> dict:
    """将嵌套推文字典展平为导出所需的扁平结构"""
    author = tweet.get("author") or {}
    if not isinstance(author, dict):
        author = {}
    metrics = tweet.get("metrics") or {}
    if not isinstance(metrics, dict):
        metrics = {}
    reply_to = tweet.get("reply_to") or {}
    if not isinstance(reply_to, dict):
        reply_to = {}
    media_list = tweet.get("media") or []

    flat = {}
    for field, _ in EXPORT_FIELDS:
        value = tweet.get(field)
        # 只有"标量"字段才能直接走短路；list/dict 必须走下面的派生逻辑做字符串化，
        # 否则 hashtags 这类 list 直接流进 CSV/Excel 会导致 openpyxl 抛
        # "Cannot convert [...] to Excel"，或 CSV 输出 "['a','b']" 这种不可读内容。
        if value is not None and not isinstance(value, (list, dict)):
            flat[field] = value
            continue

        # 从嵌套结构提取
        if field == "author_name":
            value = author.get("name", "")
        elif field == "author_username":
            value = author.get("username", "") or author.get("screen_name", "")
        elif field == "author_id":
            value = author.get("id", "")
        elif field == "author_verified":
            value = _format_verified_status(author, tweet.get("platform", "x"))
        elif field == "author_verified_type":
            value = _format_verified_type(author, tweet.get("platform", "x"))
        elif field == "author_verified_reason":
            value = author.get("verified_reason", "")
        elif field == "author_professional":
            # X: professional_type + category; 微博: mbtype/mbrank
            pro = author.get("professional_type", "")
            cat = author.get("professional_category", "")
            if pro:
                value = f"{pro}" + (f" ({cat})" if cat else "")
            elif author.get("mbrank", 0) > 0:
                value = f"微博会员 VIP{author.get('mbrank', 0)}"
            else:
                value = ""
        elif field == "author_followers":
            value = author.get("followers_count", "")
        elif field == "is_author":
            # 微博评论数据中 is_author 标识是否为原帖作者
            is_author_val = tweet.get("is_author", False)
            value = "是" if is_author_val else ""
        elif field == "like_count":
            value = tweet.get("like_count") or metrics.get("likes", "")
        elif field == "retweet_count":
            value = tweet.get("retweet_count") or metrics.get("retweets", "")
        elif field == "reply_count":
            value = tweet.get("reply_count") or metrics.get("replies", "")
        elif field == "quote_count":
            value = tweet.get("quote_count") or metrics.get("quotes", "")
        elif field == "view_count":
            value = tweet.get("view_count") or metrics.get("views", "")
        elif field == "bookmark_count":
            value = tweet.get("bookmark_count") or metrics.get("bookmarks", "")
        elif field == "reply_to_tweet_id":
            value = reply_to.get("tweet_id", "")
        elif field == "reply_to_username":
            value = reply_to.get("screen_name", "")
        elif field == "is_reply":
            value = bool(reply_to.get("tweet_id"))
        elif field == "has_media":
            value = bool(media_list or tweet.get("photos") or tweet.get("videos"))
        elif field == "hashtags":
            tags = tweet.get("hashtags", [])
            value = ", ".join(tags) if isinstance(tags, list) else str(tags) if tags else ""
        elif field == "user_mentions_text":
            mentions = tweet.get("user_mentions", [])
            if isinstance(mentions, list):
                names = [m.get("screen_name", "") for m in mentions if isinstance(m, dict)]
                value = ", ".join(n for n in names if n)
            else:
                value = ""
        elif field == "media_types":
            if isinstance(media_list, list):
                types = [m.get("type", "") for m in media_list if isinstance(m, dict)]
                value = ", ".join(t for t in types if t)
            else:
                value = ""
        elif field == "media_urls":
            if isinstance(media_list, list):
                urls = []
                for m in media_list:
                    if not isinstance(m, dict):
                        continue
                    u = m.get("video_url") or m.get("url", "")
                    if u:
                        urls.append(u)
                value = ", ".join(urls)
            else:
                value = ""
        else:
            value = ""
        flat[field] = value
    return flat


def _as_dict(value: object) -> dict:
    """将可能为 None/非 dict 的字段安全归一化为 dict。"""
    return value if isinstance(value, dict) else {}


def _count_nested_replies(replies: object) -> int:
    if not isinstance(replies, list):
        return 0

    total = 0
    stack = list(replies)
    while stack:
        node = stack.pop()
        if not isinstance(node, dict):
            continue
        total += 1
        children = node.get("replies")
        if isinstance(children, list) and children:
            stack.extend(children)
    return total


def _choose_richer_replies(*candidates: object) -> list[dict] | None:
    best: list[dict] | None = None
    best_count = -1
    for candidate in candidates:
        if not isinstance(candidate, list):
            continue
        current_count = _count_nested_replies(candidate)
        if current_count > best_count:
            best = candidate
            best_count = current_count
    return best


def _deduplicate_reply_nodes(nodes: list[dict]) -> list[dict]:
    deduped: dict[str, dict] = {}
    ordered_ids: list[str] = []
    for node in nodes:
        if not isinstance(node, dict):
            continue
        node_id = str(node.get("id") or "").strip()
        if not node_id:
            continue
        existing = deduped.get(node_id)
        if existing is None:
            deduped[node_id] = dict(node)
            ordered_ids.append(node_id)
            continue
        merged = dict(existing)
        richer = _choose_richer_replies(existing.get("replies"), node.get("replies"))
        if richer is not None:
            merged["replies"] = richer
        deduped[node_id] = merged
    return [deduped[node_id] for node_id in ordered_ids]


def _load_raw_reply_map(task_id: str, tweet_ids: list[str]) -> dict[str, list[dict]]:
    if not task_id or not tweet_ids:
        return {}

    from config import settings
    from crawler.reply_parser import parse_tweet_detail_response

    backend_dir = Path(__file__).resolve().parents[2]
    raw_root = Path(settings.raw_responses_dir)
    if not raw_root.is_absolute():
        raw_root = backend_dir / raw_root
    replies_root = raw_root / task_id / "replies"
    if not replies_root.exists():
        return {}

    reply_map: dict[str, list[dict]] = {}
    for tweet_id in tweet_ids:
        normalized_id = str(tweet_id or "").strip()
        if not normalized_id:
            continue
        reply_dir = replies_root / normalized_id
        if not reply_dir.exists():
            continue
        replies: list[dict] = []
        for reply_file in sorted(reply_dir.glob("page_*.json")):
            try:
                payload = json.loads(reply_file.read_text(encoding="utf-8"))
                _focal, parsed_replies, _bottom, _top, _has_spam_boundary = parse_tweet_detail_response(
                    payload,
                    focal_tweet_id=normalized_id,
                )
            except Exception as exc:
                logger.warning("解析原始评论响应失败: task_id=%s tweet_id=%s file=%s error=%s", task_id, normalized_id, reply_file, exc)
                continue
            replies.extend(parsed_replies)
        if replies:
            reply_map[normalized_id] = _deduplicate_reply_nodes(replies)
    return reply_map


def _build_source_reply_map(task: dict) -> dict[str, list[dict]]:
    source_ids: list[str] = []
    source_task_id = str(task.get("source_task_id") or "").strip()
    if source_task_id:
        source_ids.append(source_task_id)

    raw_source_ids = task.get("source_task_ids") or []
    if isinstance(raw_source_ids, list):
        for source_id in raw_source_ids:
            normalized = str(source_id or "").strip()
            if normalized and normalized not in source_ids:
                source_ids.append(normalized)

    if not source_ids:
        return {}

    reply_map: dict[str, list[dict]] = {}
    for source_id in source_ids:
        payload = task_manager.get_task_export_payload_readonly(source_id)
        if not payload:
            continue
        for tweet in payload.get("tweets", []):
            if not isinstance(tweet, dict):
                continue
            tweet_id = str(tweet.get("id") or "").strip()
            replies = tweet.get("replies")
            if tweet_id and isinstance(replies, list):
                reply_map[tweet_id] = replies
    return reply_map


def _hydrate_tweets_for_export(task: dict, tweets: list[dict]) -> list[dict]:
    """为导出补齐历史任务里缺失的 replies 字段。"""
    if not tweets:
        return []

    platform = str(task.get("platform") or "x").lower()
    task_kind = str(task.get("task_kind") or "search")
    replies_fetched = int(task.get("replies_fetched", 0) or 0)
    should_hydrate = bool(task.get("fetch_replies")) or replies_fetched > 0 or task_kind in {
        "comment_backfill",
        "comment_backfill_group",
    }
    if not should_hydrate:
        return list(tweets)

    source_reply_map = _build_source_reply_map(task)
    missing_ids = [
        str(tweet.get("id") or "").strip()
        for tweet in tweets
        if isinstance(tweet, dict) and tweet.get("replies") is None
    ]
    cached_reply_map = task_db.load_cached_replies_map(missing_ids) if platform == "x" else {}
    raw_reply_map = _load_raw_reply_map(str(task.get("task_id") or ""), missing_ids) if platform == "x" else {}

    # YouTube：把 raw_responses 里的原始评论响应重新解析成 replies，补给缺失的视频
    youtube_reply_map: dict[str, list[dict]] = {}
    if platform == "youtube":
        task_id_str = str(task.get("task_id") or "")
        # 对 YouTube 任务：replies 为 None 或空 list 都尝试 replay（让 _choose_richer_replies 选更多的版本）
        youtube_replay_ids = [
            str(tweet.get("id") or "").strip()
            for tweet in tweets
            if isinstance(tweet, dict)
            and not (isinstance(tweet.get("replies"), list) and len(tweet["replies"]) > 0)
        ]
        youtube_replay_ids = [vid for vid in youtube_replay_ids if vid]
        if task_id_str and youtube_replay_ids:
            try:
                from crawler.youtube import replay as yt_replay
                if yt_replay.has_raw_responses(task_id_str):
                    for vid in youtube_replay_ids:
                        rebuilt = yt_replay.rebuild_replies_for_video(task_id_str, vid)
                        if rebuilt:
                            youtube_reply_map[vid] = rebuilt
                    if youtube_reply_map:
                        logger.info(
                            "导出时从原始响应重建 YouTube 评论: task=%s videos=%s",
                            task_id_str, len(youtube_reply_map),
                        )
            except Exception as exc:
                logger.warning("YouTube 原始响应 replay 失败 task=%s: %s", task_id_str, exc)

    hydrated: list[dict] = []
    changed = False
    for tweet in tweets:
        if not isinstance(tweet, dict):
            hydrated.append(tweet)
            continue

        tweet_id = str(tweet.get("id") or "").strip()
        current_replies = tweet.get("replies")
        candidate_replies = _choose_richer_replies(
            current_replies,
            raw_reply_map.get(tweet_id),
            source_reply_map.get(tweet_id),
            cached_reply_map.get(tweet_id),
            youtube_reply_map.get(tweet_id),
        )
        if candidate_replies is current_replies or candidate_replies is None:
            hydrated.append(tweet)
            continue

        hydrated_tweet = dict(tweet)
        hydrated_tweet["replies"] = candidate_replies
        hydrated.append(hydrated_tweet)
        changed = True

    return hydrated if changed else list(tweets)


def _collect_replies(
    replies: list[dict],
    parent_id: str,
    platform: str,
    parent_author_name: str = "",
) -> list[dict]:
    """递归收集回复/评论，标注数据类型、所属推文ID、回复目标用户。

    对于一级评论（直接回复原帖的），如果本身没有 reply_to 字段，
    则自动填充 reply_to 为原帖作者。
    """
    all_replies = []
    for reply in replies:
        reply_copy = dict(reply)
        reply_copy["row_type"] = "评论"
        reply_copy["parent_tweet_id"] = parent_id
        reply_copy.setdefault("platform", platform)

        # 如果评论自身没有 reply_to 信息（一级评论）,
        # 自动填充为原帖/父级作者
        existing_reply_to = _as_dict(reply_copy.get("reply_to"))
        if not existing_reply_to.get("screen_name") and parent_author_name:
            normalized_reply_to = dict(existing_reply_to)
            normalized_reply_to["screen_name"] = parent_author_name
            reply_copy["reply_to"] = normalized_reply_to

        all_replies.append(reply_copy)

        nested = reply.get("replies", [])
        if nested and isinstance(nested, list):
            # 子评论的 parent_author_name 用当前评论的作者
            current_author = ""
            author = reply.get("author") or {}
            if isinstance(author, dict):
                current_author = author.get("name", "") or author.get("screen_name", "")
            all_replies.extend(_collect_replies(
                nested,
                parent_id=parent_id,
                platform=platform,
                parent_author_name=current_author or parent_author_name,
            ))
    return all_replies


def _collect_all_rows(tweets: list[dict], platform: str) -> list[dict]:
    """递归收集所有推文及其嵌套评论/回复，展平为独立行，并标注数据类型"""
    all_rows = []
    for tweet in tweets:
        tweet_copy = dict(tweet)
        tweet_copy["row_type"] = "原帖"
        tweet_copy["parent_tweet_id"] = ""
        tweet_copy.setdefault("platform", platform)
        all_rows.append(tweet_copy)

        # 提取原帖作者名称，供一级评论自动填充 reply_to
        author = tweet.get("author") or {}
        parent_author_name = ""
        if isinstance(author, dict):
            parent_author_name = (
                author.get("name", "")
                or author.get("screen_name", "")
            )

        replies = tweet.get("replies", [])
        if replies and isinstance(replies, list):
            all_rows.extend(_collect_replies(
                replies,
                parent_id=tweet.get("id", ""),
                platform=platform,
                parent_author_name=parent_author_name,
            ))
    return all_rows


def _make_row_dedup_key(row: dict) -> tuple[str, str, str, str, str]:
    """生成导出去重键：平台 + ID + 文本 + 数据类型 + 所属推文 ID。"""
    return (
        str(row.get("platform", "")),
        str(row.get("id", "")),
        str(row.get("text", "")),
        str(row.get("row_type", "")),
        str(row.get("parent_tweet_id", "")),
    )


def _dedup_rows(rows: list[dict]) -> list[dict]:
    """根据核心字段去重，完全相同的帖子/评论只保留第一条。"""
    seen: set[tuple[str, str, str, str, str]] = set()
    unique: list[dict] = []
    for row in rows:
        key = _make_row_dedup_key(row)
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _get_task_data(task_id: str, *, deduplicate: bool = False) -> tuple[dict, list[dict]]:
    """获取任务元信息和推文列表（含回复展平），不存在则抛 404。
    使用只读接口避免 deepcopy 开销。"""
    task = task_manager.get_task_export_payload_readonly(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务不存在: {task_id}")
    tweets = _hydrate_tweets_for_export(task, task.get("tweets", []))
    if not tweets:
        raise HTTPException(status_code=204, detail="该任务暂无数据可供导出")

    # 递归展平：将每条推文的 replies（及嵌套 replies）加入导出列表
    all_rows = _collect_all_rows(tweets, task.get("platform", "x"))

    tweet_count = len(tweets)
    reply_count = len(all_rows) - tweet_count
    logger.info(
        f"导出任务 {task_id}: {tweet_count} 条推文 + {reply_count} 条回复 = {len(all_rows)} 行"
    )
    if reply_count == 0 and task.get("fetch_replies"):
        logger.warning(
            f"任务 {task_id} 开启了回复抓取但导出回复数为 0，"
            f"请检查推文数据中 replies 字段是否存在"
        )

    if deduplicate:
        before = len(all_rows)
        all_rows = _dedup_rows(all_rows)
        removed = before - len(all_rows)
        if removed > 0:
            logger.info(f"导出去重: 移除 {removed} 条重复数据，剩余 {len(all_rows)} 行")

    return task, all_rows


def _make_filename(task: dict, ext: str) -> str:
    """根据关键词+时间生成导出文件名"""
    keyword = task.get("keyword", "export")
    # 清理关键词中不适合做文件名的字符
    clean_keyword = re.sub(r'[\\/:*?"<>|\s]+', '_', keyword).strip('_')
    # 截断过长的关键词
    if len(clean_keyword) > 50:
        clean_keyword = clean_keyword[:50]
    # 生成时间戳
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"{clean_keyword}_{now}.{ext}"


# ─── CSV 导出 ──────────────────────────────────────────────────────────────

def _build_csv(tweets: list[dict]) -> bytes:
    buf = io.StringIO()
    headers = [label for _, label in EXPORT_FIELDS]
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for tweet in tweets:
        flat = _flatten_tweet(tweet)
        row = {}
        for field, label in EXPORT_FIELDS:
            raw_value = flat[field]
            # CSV 兼容：list/dict 统一 JSON 字符串化，避免 csv 输出 Python repr（"['a', 'b']"）
            if isinstance(raw_value, (list, dict, set, tuple)):
                try:
                    row[label] = json.dumps(raw_value, ensure_ascii=False)
                except (TypeError, ValueError):
                    row[label] = str(raw_value)
            else:
                row[label] = raw_value
        writer.writerow(row)
    # UTF-8 BOM，兼容 Windows Excel 直接打开
    return "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")


@router.get("/{task_id}/csv", summary="导出推文为 CSV")
async def export_csv(
    task_id: str,
    deduplicate: bool = Query(default=False, description="是否对导出数据去重（完全相同的帖子/评论只保留一条）"),
):
    def _do():
        task, tweets = _get_task_data(task_id, deduplicate=deduplicate)
        data = _build_csv(tweets)
        filename = _make_filename(task, "csv")
        return data, filename

    loop = asyncio.get_running_loop()
    data, filename = await loop.run_in_executor(None, _do)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ─── Excel 导出 ────────────────────────────────────────────────────────────

def _build_excel(tweets: list[dict]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="缺少 openpyxl 依赖，请在后端执行: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "推文数据"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1D9BF0")  # X 品牌蓝
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # 写表头
    headers = [label for _, label in EXPORT_FIELDS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 写数据
    for row_idx, tweet in enumerate(tweets, 2):
        flat = _flatten_tweet(tweet)
        for col_idx, (field, _) in enumerate(EXPORT_FIELDS, 1):
            raw_value = flat.get(field, "")
            # 防御性转换：openpyxl 不接受 list/dict/set，遇到这类值一律 JSON 字符串化，
            # 避免任何新增字段/坏数据导致整个导出流程崩溃。
            if isinstance(raw_value, (list, dict, set, tuple)):
                try:
                    cell_value = json.dumps(raw_value, ensure_ascii=False)
                except (TypeError, ValueError):
                    cell_value = str(raw_value)
            else:
                cell_value = raw_value
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    # 自动列宽（与 EXPORT_FIELDS 一一对应）
    col_widths = [
        10,  # 平台
        20,  # 推文ID
        20,  # 对话ID
        20,  # 发布时间
        18,  # 发推客户端
        16,  # 作者昵称
        16,  # 作者账号
        18,  # 作者ID
        12,  # 认证状态
        14,  # 认证类型
        20,  # 认证说明
        16,  # 专业账号
        12,  # 作者粉丝数
        10,  # 是否楼主
        60,  # 推文内容
         8,  # 语言
        10,  # 点赞数
        10,  # 转发数
        10,  # 回复数
        10,  # 引用数
        10,  # 浏览数
        10,  # 收藏数
        50,  # 推文链接
        10,  # 数据类型
        20,  # 所属推文ID
        20,  # 回复目标推文ID
        16,  # 回复目标用户
        10,  # 是否转推
        10,  # 是否引用
        10,  # 是否回复
        20,  # 话题标签
        20,  # 提及用户
        10,  # 含媒体
        12,  # 媒体类型
        50,  # 媒体链接
    ]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/{task_id}/excel", summary="导出推文为 Excel（xlsx）")
async def export_excel(
    task_id: str,
    deduplicate: bool = Query(default=False, description="是否对导出数据去重（完全相同的帖子/评论只保留一条）"),
):
    def _do():
        task, tweets = _get_task_data(task_id, deduplicate=deduplicate)
        data = _build_excel(tweets)
        filename = _make_filename(task, "xlsx")
        return data, filename

    loop = asyncio.get_running_loop()
    data, filename = await loop.run_in_executor(None, _do)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ─── 通用接口（向后兼容）─────────────────────────────────────────────────────

@router.get("/{task_id}", summary="导出推文（format=csv|excel）")
async def export_any(
    task_id: str,
    format: Literal["csv", "excel"] = Query(default="csv", description="导出格式"),
    deduplicate: bool = Query(default=False, description="是否对导出数据去重"),
):
    if format == "excel":
        return await export_excel(task_id, deduplicate=deduplicate)
    return await export_csv(task_id, deduplicate=deduplicate)
