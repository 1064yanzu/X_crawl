"""
批量导出路由
POST /api/v1/export/batch/csv   - 批量导出为 CSV（UTF-8 BOM）
POST /api/v1/export/batch/excel - 批量导出为 Excel（xlsx，每任务一个 Sheet）
POST /api/v1/export/batch/estimate - 导出前预估行数和文件大小
"""
import asyncio
import io
import re
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from typing import Literal
from urllib.parse import quote

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from api.services import task_manager
from api.routers.export import (
    EXPORT_FIELDS,
    _flatten_tweet,
    _collect_all_rows,
    _dedup_rows,
    _make_row_dedup_key,
    _build_csv,
)

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1/export/batch", tags=["批量数据导出"])

# 导出专用线程池，避免阻塞主事件循环
_export_pool = ThreadPoolExecutor(max_workers=4, thread_name_prefix="batch-export")


class BatchExportRequest(BaseModel):
    task_ids: list[str] = Field(..., min_length=1, description="要导出的任务 ID 列表")
    merge_mode: Literal["single", "per_task"] = Field(
        default="single",
        description="合并模式: single=合并到一个文件/Sheet, per_task=每个任务独立Sheet(仅Excel)",
    )
    deduplicate: bool = Field(
        default=False,
        description="是否对导出数据去重（完全相同的帖子/评论只保留一条）",
    )


class BatchEstimateRequest(BaseModel):
    task_ids: list[str] = Field(..., min_length=1, description="要预估的任务 ID 列表")


def _load_single_task(task_id: str) -> tuple[dict, list[dict]] | None:
    """加载单个任务的导出数据（只读模式，无 deepcopy 开销）"""
    task = task_manager.get_task_export_payload_readonly(task_id)
    if not task:
        logger.warning(f"批量导出: 任务 {task_id} 不存在，跳过")
        return None
    tweets = task.get("tweets", [])
    if not tweets:
        logger.warning(f"批量导出: 任务 {task_id} 无数据，跳过")
        return None
    all_rows = _collect_all_rows(tweets, task.get("platform", "x"))
    return task, all_rows


def _get_tasks_data(task_ids: list[str]) -> list[tuple[dict, list[dict]]]:
    """批量获取任务元信息和推文列表——使用线程池并行加载多个任务"""
    with ThreadPoolExecutor(max_workers=min(len(task_ids), 4), thread_name_prefix="export-load") as pool:
        futures = {pool.submit(_load_single_task, tid): tid for tid in task_ids}
        results = []
        for future in futures:
            result = future.result()
            if result is not None:
                results.append(result)

    if not results:
        raise HTTPException(status_code=204, detail="所有选中任务均无可导出数据")
    return results


def _make_batch_filename(tasks: list[dict], ext: str) -> str:
    """批量导出文件名"""
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    if len(tasks) == 1:
        keyword = tasks[0].get("keyword", "export")
        clean = re.sub(r'[\\/:*?"<>|\s]+', '_', keyword).strip('_')[:40]
        return f"{clean}_{now}.{ext}"
    return f"batch_{len(tasks)}tasks_{now}.{ext}"


def _build_merged_csv(tasks_data: list[tuple[dict, list[dict]]]) -> bytes:
    """将多个任务的数据合并到一个 CSV"""
    all_rows = []
    for task, rows in tasks_data:
        keyword = task.get("keyword", "")
        for row in rows:
            row["_source_keyword"] = keyword
            row["_source_task_id"] = task.get("task_id", "")
        all_rows.extend(rows)
    return _build_csv_with_source(all_rows)


def _deduplicate_tasks_data(
    tasks_data: list[tuple[dict, list[dict]]],
    *,
    merge_mode: Literal["single", "per_task"],
) -> tuple[list[tuple[dict, list[dict]]], int]:
    """批量导出去重。

    - single：跨任务全局去重（适用于合并 CSV / 单 Sheet Excel）
    - per_task：每个任务各自去重（适用于每任务独立 Sheet）
    """
    if merge_mode == "per_task":
        deduped: list[tuple[dict, list[dict]]] = []
        removed_total = 0
        for task, rows in tasks_data:
            unique_rows = _dedup_rows(rows)
            removed_total += len(rows) - len(unique_rows)
            deduped.append((task, unique_rows))
        return deduped, removed_total

    seen: set[tuple[str, str, str, str, str]] = set()
    deduped = []
    removed_total = 0
    for task, rows in tasks_data:
        unique_rows: list[dict] = []
        for row in rows:
            key = _make_row_dedup_key(row)
            if key in seen:
                removed_total += 1
                continue
            seen.add(key)
            unique_rows.append(row)
        deduped.append((task, unique_rows))
    return deduped, removed_total


def _build_csv_with_source(tweets: list[dict]) -> bytes:
    """带来源列的 CSV 构建"""
    import csv

    buf = io.StringIO()
    # 增加来源列
    extra_fields = [("_source_keyword", "来源关键词"), ("_source_task_id", "来源任务ID")]
    all_fields = extra_fields + list(EXPORT_FIELDS)
    headers = [label for _, label in all_fields]
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for tweet in tweets:
        flat = _flatten_tweet(tweet)
        flat["_source_keyword"] = tweet.get("_source_keyword", "")
        flat["_source_task_id"] = tweet.get("_source_task_id", "")
        writer.writerow({label: flat.get(field, "") for field, label in all_fields})
    return "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")


def _build_batch_excel(
    tasks_data: list[tuple[dict, list[dict]]],
    merge_mode: str,
) -> bytes:
    """构建批量导出 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="缺少 openpyxl 依赖，请在后端执行: pip install openpyxl",
        )

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1D9BF0")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    extra_fields = [("_source_keyword", "来源关键词"), ("_source_task_id", "来源任务ID")]

    col_widths = [
        20,  # 来源关键词
        20,  # 来源任务ID
        10, 20, 20, 20, 18, 16, 16, 18, 10, 12, 10,
        60, 8, 10, 10, 10, 10, 10, 10, 50,
        10, 20, 20, 16, 10, 10, 10, 20, 20, 10, 12, 50,
    ]

    def _write_sheet(ws, rows: list[dict], include_source: bool = True):
        fields = (extra_fields if include_source else []) + list(EXPORT_FIELDS)
        headers = [label for _, label in fields]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_idx, tweet in enumerate(rows, 2):
            flat = _flatten_tweet(tweet)
            flat["_source_keyword"] = tweet.get("_source_keyword", "")
            flat["_source_task_id"] = tweet.get("_source_task_id", "")
            for col_idx, (field, _) in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=flat.get(field, ""))
                cell.alignment = Alignment(wrap_text=False, vertical="top")

        widths = (col_widths[:2] if include_source else []) + col_widths[2:]
        for i, width in enumerate(widths[:len(fields)], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        ws.freeze_panes = "A2"

    if merge_mode == "single":
        ws = wb.active
        ws.title = "合并数据"
        all_rows = []
        for task, rows in tasks_data:
            keyword = task.get("keyword", "")
            task_id = task.get("task_id", "")
            for row in rows:
                row["_source_keyword"] = keyword
                row["_source_task_id"] = task_id
            all_rows.extend(rows)
        _write_sheet(ws, all_rows, include_source=True)
    else:
        # per_task 模式：每个任务独立 Sheet
        wb.remove(wb.active)
        for idx, (task, rows) in enumerate(tasks_data):
            keyword = task.get("keyword", "")
            # Sheet 名称限制 31 字符，需清理
            clean_name = re.sub(r'[\\/:*?\[\]]+', '_', keyword).strip('_')[:25]
            sheet_name = f"{idx + 1}_{clean_name}" if clean_name else f"任务{idx + 1}"
            ws = wb.create_sheet(title=sheet_name)
            for row in rows:
                row["_source_keyword"] = keyword
                row["_source_task_id"] = task.get("task_id", "")
            _write_sheet(ws, rows, include_source=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.post("/estimate", summary="预估批量导出的数据量和文件大小")
async def batch_export_estimate(req: BatchEstimateRequest):
    """返回导出预估信息，帮助前端展示数据量和预估等待时间。"""
    return task_manager.get_export_estimate(req.task_ids)


@router.post("/csv", summary="批量导出为 CSV（合并到一个文件）")
async def batch_export_csv(req: BatchExportRequest):
    def _do():
        tasks_data = _get_tasks_data(req.task_ids)
        if req.deduplicate:
            td, removed = _deduplicate_tasks_data(tasks_data, merge_mode="single")
            if removed > 0:
                logger.info("批量导出 CSV 去重: 共移除 %d 条重复数据", removed)
            tasks_data = td
        data = _build_merged_csv(tasks_data)
        tasks = [t for t, _ in tasks_data]
        filename = _make_batch_filename(tasks, "csv")
        total_rows = sum(len(rows) for _, rows in tasks_data)
        logger.info(f"批量导出 CSV: {len(tasks)} 个任务, {total_rows} 行")
        return data, filename

    loop = asyncio.get_running_loop()
    data, filename = await loop.run_in_executor(_export_pool, _do)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/excel", summary="批量导出为 Excel（支持合并或分Sheet）")
async def batch_export_excel(req: BatchExportRequest):
    def _do():
        tasks_data = _get_tasks_data(req.task_ids)
        if req.deduplicate:
            td, removed = _deduplicate_tasks_data(tasks_data, merge_mode=req.merge_mode)
            if removed > 0:
                logger.info("批量导出 Excel 去重: 共移除 %d 条重复数据, mode=%s", removed, req.merge_mode)
            tasks_data = td
        data = _build_batch_excel(tasks_data, req.merge_mode)
        tasks = [t for t, _ in tasks_data]
        filename = _make_batch_filename(tasks, "xlsx")
        total_rows = sum(len(rows) for _, rows in tasks_data)
        logger.info(f"批量导出 Excel: {len(tasks)} 个任务, {total_rows} 行, mode={req.merge_mode}")
        return data, filename

    loop = asyncio.get_running_loop()
    data, filename = await loop.run_in_executor(_export_pool, _do)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
