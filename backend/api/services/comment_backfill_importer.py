from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Iterable, Literal, Optional
from urllib.parse import urlparse

from fastapi import HTTPException
from openpyxl import load_workbook

Platform = Literal["x", "weibo"]

_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "platform": ("平台", "platform"),
    "row_type": ("数据类型", "row_type"),
    "id": ("推文ID", "微博ID", "id"),
    "conversation_id": ("对话ID", "conversation_id"),
    "created_at": ("发布时间", "created_at"),
    "source": ("发推客户端", "来源", "source"),
    "author_name": ("作者昵称", "author_name"),
    "author_username": ("作者账号", "author_username"),
    "author_id": ("作者ID", "author_id"),
    "author_verified": ("认证状态", "author_verified"),
    "author_followers": ("作者粉丝数", "author_followers"),
    "text": ("推文内容", "微博内容", "text"),
    "lang": ("语言", "lang"),
    "like_count": ("点赞数", "like_count"),
    "retweet_count": ("转发数", "retweet_count"),
    "reply_count": ("回复数", "评论数", "reply_count"),
    "quote_count": ("引用数", "quote_count"),
    "view_count": ("浏览数", "view_count"),
    "bookmark_count": ("收藏数", "bookmark_count"),
    "url": ("推文链接", "微博链接", "url"),
    "hashtags": ("话题标签", "hashtags"),
    "user_mentions_text": ("提及用户", "user_mentions_text"),
    "has_media": ("含媒体", "has_media"),
    "media_types": ("媒体类型", "media_types"),
    "media_urls": ("媒体链接", "media_urls"),
}


@dataclass
class CommentBackfillImportResult:
    summary: dict
    tweets: list[dict]


def analyze_comment_backfill_file(
    filename: str,
    content: bytes,
    platform: Platform,
) -> CommentBackfillImportResult:
    rows = _load_rows(filename, content)
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件为空，无法补采评论")

    normalized_rows = [_normalize_row(row) for row in rows]
    if not normalized_rows:
        raise HTTPException(status_code=400, detail="导入文件缺少可识别的表头")

    detected_platforms = {
        value for value in (_normalize_platform(row.get("platform")) for row in normalized_rows) if value
    }
    detected_platform = next(iter(detected_platforms)) if len(detected_platforms) == 1 else None
    if detected_platform and detected_platform != platform:
        raise HTTPException(
            status_code=400,
            detail=f"导入文件平台为 {detected_platform}，与当前选择的 {platform} 不一致",
        )

    original_rows = 0
    skipped_non_post_rows = 0
    skipped_zero_comment_posts = 0
    skipped_invalid_posts = 0
    deduplicated_posts = 0
    unique_posts = 0
    tweets: list[dict] = []
    seen_ids: set[str] = set()

    for row in normalized_rows:
        row_type = str(row.get("row_type") or "").strip()
        if row_type and row_type != "原帖":
            skipped_non_post_rows += 1
            continue

        original_rows += 1
        tweet = _build_seed_tweet(row, platform)
        if tweet is None:
            skipped_invalid_posts += 1
            continue

        tweet_id = str(tweet.get("id") or "").strip()
        if tweet_id in seen_ids:
            deduplicated_posts += 1
            continue
        seen_ids.add(tweet_id)
        unique_posts += 1

        reply_count = int((tweet.get("metrics") or {}).get("replies") or 0)
        if reply_count <= 0:
            skipped_zero_comment_posts += 1
            continue

        tweets.append(tweet)

    if not tweets:
        raise HTTPException(status_code=400, detail="导入文件中没有可补采评论的原帖")

    summary = {
        "file_name": filename,
        "platform": platform,
        "total_rows": len(rows),
        "original_post_rows": original_rows,
        "unique_post_count": unique_posts,
        "eligible_posts": len(tweets),
        "skipped_non_post_rows": skipped_non_post_rows,
        "skipped_zero_comment_posts": skipped_zero_comment_posts,
        "skipped_invalid_posts": skipped_invalid_posts,
        "deduplicated_posts": deduplicated_posts,
        "has_platform_column": any("platform" in row for row in normalized_rows),
        "detected_platform": detected_platform,
    }
    return CommentBackfillImportResult(summary=summary, tweets=tweets)


def _load_rows(filename: str, content: bytes) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        return _load_csv_rows(content)
    if suffix in {"xlsx", "xlsm"}:
        return _load_excel_rows(content)
    raise HTTPException(status_code=400, detail="仅支持导入 CSV 或 XLSX 文件")


def _load_csv_rows(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _load_excel_rows(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    records: list[dict] = []
    for values in rows[1:]:
        if values is None:
            continue
        row = {}
        empty = True
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = values[idx] if idx < len(values) else None
            if value not in (None, ""):
                empty = False
            row[header] = value
        if not empty:
            records.append(row)
    return records


def _normalize_row(row: dict) -> dict:
    normalized = {}
    for field, aliases in _FIELD_ALIASES.items():
        for alias in aliases:
            if alias in row:
                normalized[field] = row.get(alias)
                break
    return normalized


def _build_seed_tweet(row: dict, platform: Platform) -> Optional[dict]:
    tweet_id = _clean_str(row.get("id"))
    if not tweet_id:
        return None

    reply_count = _to_int(row.get("reply_count"))
    url = _clean_str(row.get("url"))
    author_name = _clean_str(row.get("author_name"))
    author_username = _derive_screen_name(
        _clean_str(row.get("author_username")),
        url=url,
        platform=platform,
    )
    author_id = _clean_str(row.get("author_id"))

    if platform == "x" and not author_username:
        return None
    if platform == "weibo" and not (author_id or url):
        return None

    return {
        "id": tweet_id,
        "conversation_id": _clean_str(row.get("conversation_id")),
        "created_at": _clean_str(row.get("created_at")),
        "source": _clean_str(row.get("source")),
        "text": _clean_str(row.get("text")),
        "lang": _clean_str(row.get("lang")),
        "url": url,
        "platform": platform,
        "author": {
            "id": author_id,
            "name": author_name or author_username,
            "screen_name": author_username or author_name,
            "verified": bool(_clean_str(row.get("author_verified"))),
            "followers_count": _to_int(row.get("author_followers")),
        },
        "metrics": {
            "likes": _to_int(row.get("like_count")),
            "retweets": _to_int(row.get("retweet_count")),
            "replies": reply_count,
            "quotes": _to_int(row.get("quote_count")),
            "views": _nullable_int(row.get("view_count")),
            "bookmarks": _to_int(row.get("bookmark_count")),
        },
        "hashtags": _split_csv_like(row.get("hashtags")),
        "user_mentions": [
            {"screen_name": item}
            for item in _split_csv_like(row.get("user_mentions_text"))
        ],
        "media": _build_media(row),
        "is_retweet": _to_bool(row.get("is_retweet")),
        "is_quote": _to_bool(row.get("is_quote")),
    }


def _build_media(row: dict) -> list[dict]:
    urls = _split_csv_like(row.get("media_urls"))
    types = _split_csv_like(row.get("media_types"))
    if not urls:
        return []
    media = []
    for idx, media_url in enumerate(urls):
        media_type = types[idx] if idx < len(types) else ""
        media.append({"type": media_type or "photo", "url": media_url})
    return media


def _derive_screen_name(raw: str, *, url: str, platform: Platform) -> str:
    if raw:
        return raw.lstrip("@")
    if not url:
        return ""
    parsed = urlparse(url)
    parts = [part for part in parsed.path.split("/") if part]
    if platform == "x" and len(parts) >= 3 and parts[1] == "status":
        return parts[0].lstrip("@")
    return ""


def _normalize_platform(value: object) -> Optional[Platform]:
    text = _clean_str(value).lower()
    if text in {"x", "twitter"}:
        return "x"
    if text in {"weibo", "微博"}:
        return "weibo"
    return None


def _split_csv_like(value: object) -> list[str]:
    text = _clean_str(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[，,;；]\s*", text) if part.strip()]


def _to_bool(value: object) -> bool:
    text = _clean_str(value).lower()
    return text in {"1", "true", "yes", "y", "是"}


def _to_int(value: object) -> int:
    text = _clean_str(value)
    if not text:
        return 0
    try:
        return int(float(text.replace(",", "")))
    except ValueError:
        return 0


def _nullable_int(value: object) -> Optional[int]:
    text = _clean_str(value)
    if not text:
        return None
    try:
        return int(float(text.replace(",", "")))
    except ValueError:
        return None


def _clean_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()
